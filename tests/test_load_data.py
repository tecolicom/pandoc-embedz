"""Tests for data loading functionality"""
import pytest
from pathlib import Path
from io import StringIO
from pandoc_embedz.data_loader import load_data, guess_format_from_filename

FIXTURES_DIR = Path(__file__).parent / 'fixtures'


class TestGuessFormat:
    """Tests for format auto-detection"""

    def test_csv_extension(self):
        assert guess_format_from_filename('data.csv') == 'csv'

    def test_tsv_extension(self):
        assert guess_format_from_filename('data.tsv') == 'tsv'

    def test_json_extension(self):
        assert guess_format_from_filename('data.json') == 'json'

    def test_yaml_extension(self):
        assert guess_format_from_filename('data.yaml') == 'yaml'
        assert guess_format_from_filename('data.yml') == 'yaml'

    def test_toml_extension(self):
        assert guess_format_from_filename('data.toml') == 'toml'

    def test_sqlite_extensions(self):
        assert guess_format_from_filename('data.db') == 'sqlite'
        assert guess_format_from_filename('data.sqlite') == 'sqlite'
        assert guess_format_from_filename('data.sqlite3') == 'sqlite'

    def test_txt_extension(self):
        assert guess_format_from_filename('data.txt') == 'lines'

    def test_default_to_csv(self):
        assert guess_format_from_filename('data.unknown') == 'csv'


class TestLoadCSV:
    """Tests for CSV data loading"""

    def test_load_csv_with_header(self):
        data = load_data(str(FIXTURES_DIR / 'sample.csv'), format='csv', has_header=True)
        assert len(data) == 3
        assert data[0]['name'] == 'Arthur'
        assert data[0]['value'] == 42
        assert data[0]['category'] == 'A'

    def test_load_csv_without_header(self):
        data = load_data(str(FIXTURES_DIR / 'sample.csv'), format='csv', has_header=False)
        assert len(data) == 4  # Including header row as data
        assert isinstance(data[0], list)

    def test_load_csv_inline(self):
        csv_data = StringIO("name,value\nArthur,100\nFord,85")
        data = load_data(csv_data, format='csv', has_header=True)
        assert len(data) == 2
        assert data[0]['name'] == 'Arthur'

    def test_load_csv_with_query(self):
        """Test CSV with SQL query filter"""
        data = load_data(
            str(FIXTURES_DIR / 'sample.csv'),
            format='csv',
            query='SELECT * FROM data WHERE category = "A"'
        )
        assert len(data) == 2
        assert data[0]['name'] == 'Arthur'
        assert data[1]['name'] == 'Zaphod'

    def test_load_csv_with_aggregation_query(self):
        """Test CSV with SQL aggregation"""
        data = load_data(
            str(FIXTURES_DIR / 'sample.csv'),
            format='csv',
            query='SELECT category, COUNT(*) as count, AVG(value) as avg_value FROM data GROUP BY category'
        )
        assert len(data) == 2
        # Category A has 2 items (Arthur:42, Zaphod:99), avg = 70.5
        # Category B has 1 item (Ford:100), avg = 100
        cat_a = [row for row in data if row['category'] == 'A'][0]
        assert cat_a['count'] == 2
        assert cat_a['avg_value'] == 70.5

    def test_load_csv_with_order_by_query(self):
        """Test CSV with SQL ORDER BY"""
        data = load_data(
            str(FIXTURES_DIR / 'sample.csv'),
            format='csv',
            query='SELECT * FROM data ORDER BY value DESC'
        )
        assert len(data) == 3
        assert data[0]['name'] == 'Ford'  # value = 100
        assert data[1]['name'] == 'Zaphod'  # value = 99
        assert data[2]['name'] == 'Arthur'  # value = 42


class TestCSVComments:
    """Tests for CSV comment handling"""

    def test_csv_comment_line_default(self):
        """Default: all # lines are stripped"""
        csv_data = StringIO("# auto-generated\nname,value\n# mid comment\nAlice,100\n")
        data = load_data(csv_data, format='csv')
        assert len(data) == 1
        assert data[0]['name'] == 'Alice'

    def test_csv_comment_head(self):
        """comment=head preserves # in data rows"""
        csv_data = StringIO("# comment\nname,value\n#tag,100\n")
        data = load_data(csv_data, format='csv', comment='head')
        assert len(data) == 1
        assert data[0]['name'] == '#tag'

    def test_csv_comment_none(self):
        """comment=none disables comment handling"""
        csv_data = StringIO("# comment,x\nname,value\nAlice,100\n")
        data = load_data(csv_data, format='csv', has_header=False, comment='none')
        assert len(data) == 3
        assert data[0][0] == '# comment'

    def test_csv_comment_line(self):
        """comment=line strips all # lines"""
        csv_data = StringIO("# header comment\nname,value\n# mid comment\nAlice,100\n")
        data = load_data(csv_data, format='csv', comment='line')
        assert len(data) == 1
        assert data[0]['name'] == 'Alice'

    def test_csv_comment_inline(self):
        """comment=inline strips from # to end of line"""
        csv_data = StringIO("name,value\nAlice,100 # note\n")
        data = load_data(csv_data, format='csv', comment='inline')
        assert len(data) == 1
        assert data[0]['value'] == 100

    def test_csv_comment_false(self):
        """comment=False disables comment handling"""
        csv_data = StringIO("# data,x\nname,value\n")
        data = load_data(csv_data, format='csv', has_header=False, comment=False)
        assert len(data) == 2
        assert data[0][0] == '# data'

    def test_tsv_comment_head(self):
        """Comment handling works with TSV"""
        tsv_data = StringIO("# comment\nname\tvalue\nAlice\t100\n")
        data = load_data(tsv_data, format='tsv')
        assert len(data) == 1
        assert data[0]['name'] == 'Alice'


class TestLoadTSV:
    """Tests for TSV data loading"""

    def test_load_tsv_with_header(self):
        data = load_data(str(FIXTURES_DIR / 'sample.tsv'), format='tsv', has_header=True)
        assert len(data) == 3
        assert data[0]['name'] == 'Arthur'
        assert data[1]['name'] == 'Ford'

    def test_load_tsv_inline(self):
        tsv_data = StringIO("name\tvalue\nArthur\t42\nFord\t100")
        data = load_data(tsv_data, format='tsv', has_header=True)
        assert len(data) == 2
        assert data[0]['value'] == 42

    def test_load_tsv_with_spaces_in_values(self):
        """Test that TSV correctly handles spaces within field values"""
        tsv_data = StringIO("name\tvalue\tcomment\nArthur Dent\t42\tHoopy frood\nFord Prefect\t100\tGreat guy")
        data = load_data(tsv_data, format='tsv', has_header=True)

        assert len(data) == 2
        # Spaces within values should be preserved
        assert data[0]['name'] == 'Arthur Dent'
        assert data[0]['comment'] == 'Hoopy frood'
        assert data[1]['name'] == 'Ford Prefect'
        assert data[1]['comment'] == 'Great guy'

    def test_load_tsv_without_header(self):
        """Test TSV without header row"""
        tsv_data = StringIO("Arthur Dent\t42\nFord Prefect\t100")
        data = load_data(tsv_data, format='tsv', has_header=False)

        assert len(data) == 2
        assert data[0][0] == 'Arthur Dent'
        assert data[0][1] == 42


class TestLoadSSV:
    """Tests for SSV (space-separated) data loading"""

    def test_load_ssv_with_header(self):
        ssv_data = StringIO("name value category\nArthur 42 A\nFord 100 B")
        data = load_data(ssv_data, format='ssv', has_header=True)
        assert len(data) == 2
        assert data[0]['name'] == 'Arthur'
        assert data[0]['value'] == 42

    def test_load_ssv_without_header(self):
        ssv_data = StringIO("Arthur 42 A\nFord 100 B")
        data = load_data(ssv_data, format='ssv', has_header=False)
        assert len(data) == 2
        assert isinstance(data[0], list)
        assert data[0][0] == 'Arthur'

    def test_load_spaces_alias_with_header(self):
        """Test that 'spaces' is an alias for 'ssv' with header"""
        spaces_data = StringIO("name value category\nArthur 42 A\nFord 100 B")
        data = load_data(spaces_data, format='spaces', has_header=True)
        assert len(data) == 2
        assert data[0]['name'] == 'Arthur'
        assert data[0]['value'] == 42

    def test_load_spaces_alias_without_header(self):
        """Test that 'spaces' is an alias for 'ssv' without header"""
        spaces_data = StringIO("Arthur 42 A\nFord 100 B")
        data = load_data(spaces_data, format='spaces', has_header=False)
        assert len(data) == 2
        assert isinstance(data[0], list)
        assert data[0][0] == 'Arthur'

    def test_load_ssv_ragged_rows_strip_nan(self):
        """Test that ragged rows (unequal column count) don't produce NaN"""
        ssv_data = StringIO("a b c d e f\ng h i j\nk l")
        data = load_data(ssv_data, format='ssv', has_header=False)
        assert len(data) == 3
        assert data[0] == ['a', 'b', 'c', 'd', 'e', 'f']
        assert data[1] == ['g', 'h', 'i', 'j']
        assert data[2] == ['k', 'l']

    def test_load_csv_ragged_rows_strip_nan(self):
        """Test that ragged CSV rows don't produce NaN"""
        csv_data = StringIO("a,b,c\nd,e\nf")
        data = load_data(csv_data, format='csv', has_header=False)
        assert len(data) == 3
        assert data[0] == ['a', 'b', 'c']
        assert data[1] == ['d', 'e']
        assert data[2] == ['f']

    def test_load_ssv_with_columns_preserves_spaces_in_last_column(self):
        """Test that columns parameter preserves spaces in last column"""
        ssv_data = StringIO("ID Name Description\n1 Alice Software engineer\n2 Bob Project manager with team")
        data = load_data(ssv_data, format='ssv', has_header=True, columns=3)
        assert len(data) == 2
        assert data[0]['ID'] == '1'
        assert data[0]['Name'] == 'Alice'
        assert data[0]['Description'] == 'Software engineer'
        assert data[1]['ID'] == '2'
        assert data[1]['Name'] == 'Bob'
        assert data[1]['Description'] == 'Project manager with team'

    def test_load_ssv_with_columns_without_header(self):
        """Test columns parameter without header row"""
        ssv_data = StringIO("1 Alice Software engineer\n2 Bob Project manager with team")
        data = load_data(ssv_data, format='ssv', has_header=False, columns=3)
        assert len(data) == 2
        assert data[0] == ['1', 'Alice', 'Software engineer']
        assert data[1] == ['2', 'Bob', 'Project manager with team']

    def test_load_ssv_with_columns_multiple_spaces(self):
        """Test columns parameter handles multiple consecutive spaces"""
        ssv_data = StringIO("ID  Name  Description\n1   Alice   Has   multiple   spaces")
        data = load_data(ssv_data, format='ssv', has_header=True, columns=3)
        assert len(data) == 1
        assert data[0]['ID'] == '1'
        assert data[0]['Name'] == 'Alice'
        assert data[0]['Description'] == 'Has   multiple   spaces'

    def test_load_ssv_with_columns_fewer_fields(self):
        """Test columns parameter pads with empty strings when fewer fields"""
        ssv_data = StringIO("ID Name Description\n1 Alice\n2 Bob Complete")
        data = load_data(ssv_data, format='ssv', has_header=True, columns=3)
        assert len(data) == 2
        assert data[0]['ID'] == '1'
        assert data[0]['Name'] == 'Alice'
        assert data[0]['Description'] == ''
        assert data[1]['Description'] == 'Complete'

    def test_load_ssv_with_columns_empty_input(self):
        """Test columns parameter with empty input"""
        ssv_data = StringIO("")
        data = load_data(ssv_data, format='ssv', has_header=True, columns=3)
        assert data == []

    def test_load_ssv_with_columns_skips_blank_lines(self):
        """Test columns parameter skips blank lines"""
        ssv_data = StringIO("ID Name Description\n\n1 Alice Engineer\n\n2 Bob Manager")
        data = load_data(ssv_data, format='ssv', has_header=True, columns=3)
        assert len(data) == 2
        assert data[0]['Name'] == 'Alice'
        assert data[1]['Name'] == 'Bob'

    def test_load_ssv_with_columns_and_query(self):
        """Test columns parameter works with SQL query"""
        ssv_data = StringIO("ID Name Role\n1 Alice Engineer\n2 Bob Manager\n3 Carol Engineer")
        data = load_data(
            ssv_data,
            format='ssv',
            has_header=True,
            columns=3,
            query="SELECT * FROM data WHERE Role = 'Engineer'"
        )
        assert len(data) == 2
        assert data[0]['Name'] == 'Alice'
        assert data[1]['Name'] == 'Carol'

    def test_load_spaces_alias_with_columns(self):
        """Test that 'spaces' alias works with columns parameter"""
        spaces_data = StringIO("ID Name Description\n1 Alice Software engineer\n2 Bob Project manager")
        data = load_data(spaces_data, format='spaces', has_header=True, columns=3)
        assert len(data) == 2
        assert data[0]['Description'] == 'Software engineer'
        assert data[1]['Description'] == 'Project manager'


class TestLoadJSON:
    """Tests for JSON data loading"""

    def test_load_json_file(self):
        data = load_data(str(FIXTURES_DIR / 'sample.json'), format='json')
        assert len(data) == 3
        assert data[0]['name'] == 'Arthur'
        assert data[0]['value'] == 42

    def test_load_json_inline(self):
        json_data = StringIO('[{"name": "Arthur", "value": 42}]')
        data = load_data(json_data, format='json')
        assert len(data) == 1
        assert data[0]['name'] == 'Arthur'

    def test_load_json_object(self):
        json_data = StringIO('{"title": "Test", "count": 5}')
        data = load_data(json_data, format='json')
        assert data['title'] == 'Test'
        assert data['count'] == 5


class TestLoadYAML:
    """Tests for YAML data loading"""

    def test_load_yaml_file(self):
        data = load_data(str(FIXTURES_DIR / 'sample.yaml'), format='yaml')
        assert len(data) == 3
        assert data[0]['name'] == 'Arthur'
        assert data[0]['value'] == 42

    def test_load_yaml_inline(self):
        yaml_data = StringIO("- name: Arthur\n  value: 42\n- name: Ford\n  value: 100")
        data = load_data(yaml_data, format='yaml')
        assert len(data) == 2
        assert data[0]['name'] == 'Arthur'


class TestLoadTOML:
    """Tests for TOML data loading"""

    def test_load_toml_file(self):
        data = load_data(str(FIXTURES_DIR / 'sample.toml'), format='toml')
        assert 'items' in data
        assert len(data['items']) == 3
        assert data['items'][0]['name'] == 'Arthur'
        assert data['items'][0]['value'] == 42

    def test_load_toml_inline(self):
        toml_data = StringIO('[items]\nname = "Arthur"\nvalue = 42\n\n[config]\ntitle = "Test"')
        data = load_data(toml_data, format='toml')
        assert data['items']['name'] == 'Arthur'
        assert data['items']['value'] == 42
        assert data['config']['title'] == 'Test'


class TestLoadSQLite:
    """Tests for SQLite database loading"""

    def test_load_sqlite_with_table(self):
        data = load_data(str(FIXTURES_DIR / 'sample.db'), format='sqlite', table='items')
        assert len(data) == 3
        assert data[0]['name'] == 'Arthur'
        assert data[0]['value'] == 42
        assert data[0]['category'] == 'A'

    def test_load_sqlite_with_query(self):
        data = load_data(
            str(FIXTURES_DIR / 'sample.db'),
            format='sqlite',
            query='SELECT * FROM items WHERE category = "A"'
        )
        assert len(data) == 2
        assert data[0]['name'] == 'Arthur'
        assert data[1]['name'] == 'Zaphod'

    def test_load_sqlite_query_overrides_table(self):
        """When both query and table are specified, query takes precedence"""
        data = load_data(
            str(FIXTURES_DIR / 'sample.db'),
            format='sqlite',
            table='items',
            query='SELECT * FROM metadata'
        )
        assert len(data) == 2
        assert data[0]['key'] == 'author'

    def test_load_sqlite_requires_table_or_query(self):
        """SQLite format requires either table or query parameter"""
        with pytest.raises(ValueError, match="requires either 'table' or 'query'"):
            load_data(str(FIXTURES_DIR / 'sample.db'), format='sqlite')

    def test_load_sqlite_no_inline_support(self):
        """SQLite does not support inline data"""
        with pytest.raises(ValueError, match="does not support inline data"):
            load_data(StringIO("dummy"), format='sqlite', table='items')

    def test_load_sqlite_sql_injection_prevented(self):
        """SQL injection via table name should be safely escaped"""
        # Attempt SQL injection via table name - should be escaped, not executed
        # The table doesn't exist, but the point is it doesn't execute as SQL
        from pandoc_embedz.data_loader import _quote_identifier
        # Verify quoting works correctly
        assert _quote_identifier('items') == '"items"'
        assert _quote_identifier('items; DROP TABLE items;--') == '"items; DROP TABLE items;--"'
        assert _quote_identifier('my table') == '"my table"'
        assert _quote_identifier('test"quote') == '"test""quote"'

    def test_load_sqlite_table_name_with_special_chars(self):
        """Table names with special characters should be properly quoted"""
        from pandoc_embedz.data_loader import _quote_identifier
        # These would be dangerous without quoting, but are safe when quoted
        assert _quote_identifier('123table') == '"123table"'
        assert _quote_identifier('my-table') == '"my-table"'
        assert _quote_identifier('schema.table') == '"schema.table"'


class TestLoadLines:
    """Tests for lines format data loading"""

    def test_load_lines_file(self):
        data = load_data(str(FIXTURES_DIR / 'sample.txt'), format='lines')
        assert len(data) == 3
        assert data[0] == 'Arthur'
        assert data[1] == 'Ford'
        assert data[2] == 'Zaphod'

    def test_load_lines_inline(self):
        lines_data = StringIO("Arthur\nFord\nZaphod\n")
        data = load_data(lines_data, format='lines')
        assert len(data) == 3
        assert data[0] == 'Arthur'

    def test_load_lines_skip_empty(self):
        lines_data = StringIO("Arthur\n\nFord\n\n\nZaphod")
        data = load_data(lines_data, format='lines')
        assert len(data) == 6  # Empty lines are preserved
        assert data == ['Arthur', '', 'Ford', '', '', 'Zaphod']


class TestAutoDetection:
    """Tests for format auto-detection"""

    def test_csv_auto_detect(self):
        data = load_data(str(FIXTURES_DIR / 'sample.csv'))
        assert len(data) == 3
        assert data[0]['name'] == 'Arthur'

    def test_json_auto_detect(self):
        data = load_data(str(FIXTURES_DIR / 'sample.json'))
        assert len(data) == 3
        assert data[0]['name'] == 'Arthur'

    def test_yaml_auto_detect(self):
        data = load_data(str(FIXTURES_DIR / 'sample.yaml'))
        assert len(data) == 3
        assert data[0]['name'] == 'Arthur'

    def test_toml_auto_detect(self):
        data = load_data(str(FIXTURES_DIR / 'sample.toml'))
        assert 'items' in data
        assert len(data['items']) == 3
        assert data['items'][0]['name'] == 'Arthur'

    def test_sqlite_auto_detect(self):
        """SQLite auto-detects format but still requires table parameter"""
        data = load_data(str(FIXTURES_DIR / 'sample.db'), table='items')
        assert len(data) == 3
        assert data[0]['name'] == 'Arthur'

    def test_txt_auto_detect(self):
        data = load_data(str(FIXTURES_DIR / 'sample.txt'))
        assert len(data) == 3
        assert data[0] == 'Arthur'


class TestMultiTableSQL:
    """Tests for multi-table SQL queries (integration test via process_embedz)"""

    def test_multi_table_join(self):
        """Test joining two CSV files with SQL"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        # Clear state
        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        # Create embedz code block with multi-table data
        code = '''---
data:
  products: tests/fixtures/products.csv
  sales: tests/fixtures/sales.csv
query: |
  SELECT
    p.product_name,
    s.quantity,
    s.date
  FROM sales s
  JOIN products p ON s.product_id = p.product_id
  ORDER BY s.date
---
{% for row in data %}
- {{ row.product_name }}: {{ row.quantity }} units on {{ row.date }}
{% endfor %}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()

        result = process_embedz(elem, doc)

        # Convert result to markdown
        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        # Verify output contains joined data
        assert 'Widget' in markdown
        assert 'Gadget' in markdown
        assert 'Doohickey' in markdown
        assert '5 units' in markdown
        assert '2024-01-15' in markdown

    def test_multi_table_aggregation(self):
        """Test aggregating data from multiple tables"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        # Clear state
        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        # Create embedz code block with aggregation
        code = '''---
data:
  products: tests/fixtures/products.csv
  sales: tests/fixtures/sales.csv
query: |
  SELECT
    p.product_name,
    SUM(s.quantity) as total_quantity,
    SUM(s.quantity * p.price) as total_revenue
  FROM sales s
  JOIN products p ON s.product_id = p.product_id
  GROUP BY p.product_name
  ORDER BY total_revenue DESC
---
| Product | Quantity | Revenue |
|---------|----------|---------|
{% for row in data -%}
| {{ row.product_name }} | {{ row.total_quantity }} | ${{ "%.2f" | format(row.total_revenue) }} |
{% endfor -%}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()

        result = process_embedz(elem, doc)

        # Convert result to markdown
        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        # Verify output contains aggregated data
        assert 'Widget' in markdown
        assert 'Gadget' in markdown
        assert 'Doohickey' in markdown

    def test_multi_table_without_query(self):
        """Multi-table data without query allows direct access via data.table_name"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        # Clear state
        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        # Create embedz code block without query - accessing via data.table_name
        code = '''---
data:
  products: tests/fixtures/products.csv
  sales: tests/fixtures/sales.csv
---
## Products
{% for p in data.products %}
- {{ p.product_name }}: ¥{{ p.price }}
{% endfor %}

## Sales
{% for s in data.sales %}
- Sale #{{ s.sale_id }}: {{ s.quantity }} units
{% endfor %}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()

        result = process_embedz(elem, doc)

        # Convert result to markdown
        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        # Verify output contains data from both files
        assert 'Products' in markdown
        assert 'Sales' in markdown
        assert 'Widget' in markdown
        assert 'Gadget' in markdown
        assert 'Sale #101' in markdown
        assert 'Sale #102' in markdown

    def test_multi_table_mixed_formats(self):
        """Multi-table can combine different formats (YAML + CSV)"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        # Clear state
        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        # Create embedz code block combining YAML config and CSV data
        code = '''---
data:
  config: tests/fixtures/config.yaml
  sales: tests/fixtures/sales.csv
---
# {{ data.config.title }}
## {{ data.config.subtitle }}

By {{ data.config.author }} (v{{ data.config.version }})

{% for sale in data.sales[:3] %}
- Sale #{{ sale.sale_id }}: {{ sale.quantity }} units
{% endfor %}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()

        result = process_embedz(elem, doc)

        # Convert result to markdown
        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        # Verify output contains both config and data
        assert '2024 Sales Report' in markdown
        assert 'Q1 Results' in markdown
        assert 'John Doe' in markdown
        assert 'v1.0' in markdown
        assert 'Sale #101' in markdown

    def test_multi_table_inline_csv(self):
        """Multi-table with inline CSV data"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        # Clear state
        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        # Create embedz code block with inline CSV data
        code = '''---
data:
  products: |
    product_id,product_name,price
    1,Widget,1280
    2,Gadget,2480
  sales: |
    sale_id,product_id,quantity
    101,1,5
    102,2,3
---
## Products
{% for p in data.products %}
- {{ p.product_name }}: ¥{{ "{:,}".format(p.price|int) }}
{% endfor %}

## Sales
{% for s in data.sales %}
- Sale #{{ s.sale_id }}: {{ s.quantity }} units
{% endfor %}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()

        result = process_embedz(elem, doc)

        # Convert result to markdown
        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        # Verify output contains data from inline sources
        assert 'Products' in markdown
        assert 'Widget: ¥1,280' in markdown
        assert 'Gadget: ¥2,480' in markdown
        assert 'Sale #101: 5 units' in markdown
        assert 'Sale #102: 3 units' in markdown

    def test_multi_table_inline_yaml(self):
        """Multi-table with inline YAML config and CSV data"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        # Clear state
        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        # Create embedz code block with mixed inline data
        code = '''---
data:
  config:
    format: yaml
    data: |
      title: "Test Report"
      year: 2024
  sales: |
    date,amount
    2024-01-01,100
    2024-01-02,200
---
# {{ data.config.title }} ({{ data.config.year }})

{% for s in data.sales %}
- {{ s.date }}: ${{ s.amount }}
{% endfor %}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()

        result = process_embedz(elem, doc)

        # Convert result to markdown
        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        # Verify output contains both config and data
        assert 'Test Report (2024)' in markdown
        assert '2024-01-01: \\$100' in markdown or '2024-01-01: $100' in markdown
        assert '2024-01-02: \\$200' in markdown or '2024-01-02: $200' in markdown

    def test_multi_table_mixed_inline_and_file(self):
        """Multi-table with both inline data and file paths"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        # Clear state
        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        # Create embedz code block mixing inline and file data
        code = '''---
data:
  config:
    format: yaml
    data: |
      title: "Mixed Source Report"
  products: tests/fixtures/products.csv
  sales: |
    sale_id,product_id,quantity
    999,1,10
---
# {{ data.config.title }}

## Products from file
{% for p in data.products[:2] %}
- {{ p.product_name }}
{% endfor %}

## Sales from inline
{% for s in data.sales %}
- Sale #{{ s.sale_id }}: {{ s.quantity }} units
{% endfor %}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()

        result = process_embedz(elem, doc)

        # Convert result to markdown
        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        # Verify output contains data from both sources
        assert 'Mixed Source Report' in markdown
        assert 'Widget' in markdown  # From file
        assert 'Sale #999: 10 units' in markdown  # From inline

    def test_data_file_and_data_part_mutually_exclusive(self):
        """Error should be raised if both data attribute and inline data are specified"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        # Clear state
        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        # Create embedz code block with both data attribute and inline data
        code = '''---
data: tests/fixtures/products.csv
---
{% for p in data %}
- {{ p.product_name }}
{% endfor %}
---
product_id,product_name,price
1,Widget,1280
2,Gadget,2480'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()

        # Should raise ValueError
        import pytest
        with pytest.raises(ValueError, match="Cannot specify both 'data' attribute and inline data"):
            process_embedz(elem, doc)

class TestEmptyInput:
    """Tests for empty input handling"""

    def test_empty_json_returns_empty_list(self):
        """Empty JSON input should return empty list instead of error"""
        data = load_data(StringIO(""), format='json')
        assert data == []

    def test_whitespace_json_returns_empty_list(self):
        """Whitespace-only JSON input should return empty list"""
        data = load_data(StringIO("   \n  \t  "), format='json')
        assert data == []

    def test_empty_csv_returns_empty_list(self):
        """Empty CSV input should return empty list instead of error"""
        data = load_data(StringIO(""), format='csv')
        assert data == []

    def test_whitespace_csv_returns_empty_list(self):
        """Whitespace-only CSV input should return empty list"""
        data = load_data(StringIO("   \n  \t  "), format='csv')
        assert data == []

    def test_empty_tsv_returns_empty_list(self):
        """Empty TSV input should return empty list"""
        data = load_data(StringIO(""), format='tsv')
        assert data == []


class TestLoadExcel:
    """Tests for Excel data loading"""

    @pytest.fixture(autouse=True)
    def check_openpyxl(self):
        pytest.importorskip("openpyxl")

    def test_excel_extensions(self):
        """Extension auto-detection for .xlsx and .xls"""
        assert guess_format_from_filename('data.xlsx') == 'excel'
        assert guess_format_from_filename('data.xls') == 'excel'

    def test_load_excel_with_header(self):
        """Load Excel with header row (default)"""
        data = load_data(str(FIXTURES_DIR / 'sample.xlsx'), format='excel', table='items')
        assert len(data) == 3
        assert data[0]['name'] == 'Arthur'
        assert data[0]['value'] == 42
        assert data[0]['category'] == 'A'

    def test_load_excel_without_header(self):
        """Load Excel without header row"""
        data = load_data(str(FIXTURES_DIR / 'sample.xlsx'), format='excel', table='items', has_header=False)
        assert len(data) == 4  # Including header row as data
        assert isinstance(data[0], list)
        assert data[0][0] == 'name'  # Header row treated as data

    def test_load_excel_with_sheet_name(self):
        """Load specific sheet by name via table parameter"""
        data = load_data(str(FIXTURES_DIR / 'sample.xlsx'), format='excel', table='metadata')
        assert len(data) == 2
        assert data[0]['key'] == 'author'
        assert data[0]['description'] == 'Douglas Adams'

    def test_load_excel_default_sheet(self):
        """Load first sheet when table is not specified"""
        data = load_data(str(FIXTURES_DIR / 'sample.xlsx'), format='excel')
        assert len(data) == 3
        assert data[0]['name'] == 'Arthur'

    def test_load_excel_with_query(self):
        """Apply SQL query to Excel data"""
        data = load_data(
            str(FIXTURES_DIR / 'sample.xlsx'),
            format='excel',
            table='items',
            query='SELECT * FROM data WHERE category = "A"'
        )
        assert len(data) == 2
        assert data[0]['name'] == 'Arthur'
        assert data[1]['name'] == 'Zaphod'

    def test_load_excel_no_inline(self):
        """Excel does not support inline data"""
        with pytest.raises(ValueError, match="does not support inline data"):
            load_data(StringIO("dummy"), format='excel')

    def test_load_excel_no_openpyxl(self, monkeypatch):
        """Error message when openpyxl is not installed"""
        import pandoc_embedz.data_loader as dl
        monkeypatch.setattr(dl, 'openpyxl', None)
        with pytest.raises(ImportError, match="openpyxl"):
            load_data(str(FIXTURES_DIR / 'sample.xlsx'), format='excel')

    def test_load_excel_skips_blank_rows_and_columns(self, tmp_path):
        """Leading blank rows and all-blank columns are skipped"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([])                              # blank row
        ws.append([])                              # blank row
        ws.append([None, 'name', 'value', None])   # header with blank cols
        ws.append([None, 'Arthur', 42, None])
        ws.append([None, 'Ford', 100, None])
        path = str(tmp_path / 'blank.xlsx')
        wb.save(path)

        data = load_data(path, format='excel')
        assert len(data) == 2
        assert data[0] == {'name': 'Arthur', 'value': 42}
        assert data[1] == {'name': 'Ford', 'value': 100}

    def test_load_excel_skips_blank_rows_without_header(self, tmp_path):
        """Blank rows/columns skipped in headerless mode too"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([])
        ws.append([None, 'Arthur', 42, None])
        ws.append([None, 'Ford', 100, None])
        path = str(tmp_path / 'blank_noheader.xlsx')
        wb.save(path)

        data = load_data(path, format='excel', has_header=False)
        assert len(data) == 2
        assert data[0] == ['Arthur', 42]
        assert data[1] == ['Ford', 100]

    def test_load_excel_empty_sheet(self, tmp_path, capsys):
        """Empty sheet returns empty list with warning"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        # Don't add any data
        path = str(tmp_path / 'empty.xlsx')
        wb.save(path)

        assert load_data(path, format='excel') == []
        assert 'contains no data' in capsys.readouterr().err

        assert load_data(path, format='excel', has_header=False) == []

    def test_load_excel_transpose_with_header(self, tmp_path):
        """Transpose swaps rows and columns, first column becomes header"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['name', 'Arthur', 'Ford', 'Zaphod'])
        ws.append(['value', 42, 100, 99])
        ws.append(['category', 'A', 'B', 'A'])
        path = str(tmp_path / 'transposed.xlsx')
        wb.save(path)

        data = load_data(path, format='excel', transpose=True)
        assert len(data) == 3
        assert data[0] == {'name': 'Arthur', 'value': 42, 'category': 'A'}
        assert data[1] == {'name': 'Ford', 'value': 100, 'category': 'B'}
        assert data[2] == {'name': 'Zaphod', 'value': 99, 'category': 'A'}

    def test_load_excel_transpose_without_header(self, tmp_path):
        """Transpose without header returns list of lists"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Arthur', 'Ford'])
        ws.append([42, 100])
        path = str(tmp_path / 'transposed_noheader.xlsx')
        wb.save(path)

        data = load_data(path, format='excel', transpose=True, has_header=False)
        assert len(data) == 2
        assert data[0] == ['Arthur', 42]
        assert data[1] == ['Ford', 100]

    def test_load_excel_transpose_with_query(self, tmp_path):
        """Transpose works with SQL query"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['name', 'Arthur', 'Ford', 'Zaphod'])
        ws.append(['value', 42, 100, 99])
        path = str(tmp_path / 'transposed_query.xlsx')
        wb.save(path)

        data = load_data(
            path, format='excel', transpose=True,
            query='SELECT * FROM data WHERE value >= 99'
        )
        assert len(data) == 2
        assert data[0]['name'] == 'Ford'
        assert data[1]['name'] == 'Zaphod'

    def test_load_excel_skiprows(self, tmp_path):
        """Skip leading description rows"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Annual Report 2024'])
        ws.append(['Generated: 2024-04-01'])
        ws.append([])
        ws.append(['name', 'value'])
        ws.append(['Arthur', 42])
        ws.append(['Ford', 100])
        path = str(tmp_path / 'with_title.xlsx')
        wb.save(path)

        data = load_data(path, format='excel', skiprows=2)
        assert len(data) == 2
        assert data[0] == {'name': 'Arthur', 'value': 42}
        assert data[1] == {'name': 'Ford', 'value': 100}

    def test_load_excel_skiprows_pattern_any_column(self, tmp_path):
        """Skip rows by pattern matching any column"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Title'])
        ws.append(['Description'])
        ws.append([])
        ws.append(['name', 'value'])
        ws.append(['Arthur', 42])
        path = str(tmp_path / 'pattern.xlsx')
        wb.save(path)

        data = load_data(path, format='excel', skiprows='name')
        assert len(data) == 1
        assert data[0] == {'name': 'Arthur', 'value': 42}

    def test_load_excel_skiprows_pattern_specific_column(self, tmp_path):
        """Skip rows by pattern matching a specific column"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['info', 'Title'])
        ws.append([])
        ws.append(['name', 'value'])
        ws.append(['Arthur', 42])
        path = str(tmp_path / 'col_pattern.xlsx')
        wb.save(path)

        # Column 2 contains "value" in the header row
        data = load_data(path, format='excel', skiprows='2:value')
        assert len(data) == 1
        assert data[0] == {'name': 'Arthur', 'value': 42}

    def test_load_excel_skiprows_pattern_list(self, tmp_path):
        """Skip rows by list of patterns (all must match)"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Title', 'Report'])
        ws.append(['年', '月', '件数'])
        ws.append([2024, 4, 100])
        ws.append([2024, 5, 200])
        path = str(tmp_path / 'list_pattern.xlsx')
        wb.save(path)

        data = load_data(path, format='excel', skiprows=['年', '月'])
        assert len(data) == 2
        assert data[0] == {'年': 2024, '月': 4, '件数': 100}
        assert data[1] == {'年': 2024, '月': 5, '件数': 200}

    def test_load_excel_skiprows_pattern_list_with_column(self, tmp_path):
        """Skip rows by list with N:text format"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Header', 'Info'])
        ws.append(['年', '月', '件数'])
        ws.append([2024, 4, 100])
        path = str(tmp_path / 'list_col.xlsx')
        wb.save(path)

        data = load_data(path, format='excel', skiprows=['1:年', '2:月'])
        assert len(data) == 1
        assert data[0] == {'年': 2024, '月': 4, '件数': 100}

    def test_load_excel_skiprows_pattern_list_not_found(self, tmp_path):
        """Error when not all list patterns found in same row"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['年', '件数'])
        ws.append([2024, 100])
        path = str(tmp_path / 'list_notfound.xlsx')
        wb.save(path)

        with pytest.raises(ValueError, match="not found"):
            load_data(path, format='excel', skiprows=['年', '月'])

    def test_load_excel_skiprows_single_element_list(self, tmp_path):
        """Single-element list behaves like a string pattern"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Title'])
        ws.append(['name', 'value'])
        ws.append(['Arthur', 42])
        path = str(tmp_path / 'single_list.xlsx')
        wb.save(path)

        data = load_data(path, format='excel', skiprows=['name'])
        assert len(data) == 1
        assert data[0] == {'name': 'Arthur', 'value': 42}

    def test_load_excel_skiprows_pattern_not_found(self, tmp_path):
        """Error when skiprows pattern not found"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['name', 'value'])
        ws.append(['Arthur', 42])
        path = str(tmp_path / 'no_match.xlsx')
        wb.save(path)

        with pytest.raises(ValueError, match="not found"):
            load_data(path, format='excel', skiprows='nonexistent')

    def test_load_excel_empty_header_column(self, tmp_path):
        """Empty first header cell gets replaced with column_N"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, '4月', '5月', '6月'])
        ws.append(['件数', 10, 20, 30])
        ws.append(['割合', 0.1, 0.2, 0.3])
        path = str(tmp_path / 'empty_header.xlsx')
        wb.save(path)

        data = load_data(path, format='excel')
        assert len(data) == 2
        assert data[0]['column_0'] == '件数'
        assert data[0]['4月'] == 10
        assert data[1]['column_0'] == '割合'
        assert data[1]['6月'] == 0.3

    def test_load_excel_duplicate_column_names(self, tmp_path):
        """Duplicate column names get numeric suffix"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['name', 'value', 'value'])
        ws.append(['Arthur', 42, 100])
        path = str(tmp_path / 'dup_cols.xlsx')
        wb.save(path)

        data = load_data(path, format='excel')
        assert len(data) == 1
        assert data[0]['name'] == 'Arthur'
        assert data[0]['value'] == 42
        assert data[0]['value_1'] == 100

    def test_load_excel_whitespace_column_names(self, tmp_path):
        """Column names with leading/trailing whitespace are stripped"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([' name ', ' value '])
        ws.append(['Arthur', 42])
        path = str(tmp_path / 'ws_cols.xlsx')
        wb.save(path)

        data = load_data(path, format='excel')
        assert data[0]['name'] == 'Arthur'
        assert data[0]['value'] == 42

    def test_load_excel_nan_in_data_without_header(self, tmp_path):
        """NaN values in headerless data become empty strings"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, '4月', '5月'])
        ws.append(['件数', 10, None])
        path = str(tmp_path / 'nan_data.xlsx')
        wb.save(path)

        data = load_data(path, format='excel', has_header=False)
        assert data[0] == ['', '4月', '5月']
        assert data[1] == ['件数', 10, '']

    def test_excel_auto_detect(self):
        """Format auto-detection for .xlsx files"""
        data = load_data(str(FIXTURES_DIR / 'sample.xlsx'))
        assert len(data) == 3
        assert data[0]['name'] == 'Arthur'


class TestNormalizeDataSource:
    """Tests for _normalize_data_source 3-tuple return"""

    def test_string_path_returns_3_tuple(self):
        """String path returns (source, format, {})"""
        from pandoc_embedz.data_loader import _normalize_data_source
        source, fmt, kwargs = _normalize_data_source('data.csv', 'test')
        assert source == 'data.csv'
        assert fmt == 'csv'
        assert kwargs == {}

    def test_inline_data_returns_3_tuple(self):
        """Inline data dict returns (StringIO, format, {})"""
        from pandoc_embedz.data_loader import _normalize_data_source
        source, fmt, kwargs = _normalize_data_source(
            {'data': 'a,b\n1,2'}, 'test'
        )
        assert isinstance(source, StringIO)
        assert fmt == 'csv'
        assert kwargs == {}

    def test_file_dict_returns_3_tuple(self):
        """file: dict returns (path, format, extra_kwargs)"""
        from pandoc_embedz.data_loader import _normalize_data_source
        source, fmt, kwargs = _normalize_data_source(
            {'file': 'data/test.xlsx', 'table': 'Sheet1', 'skiprows': 'Year'},
            'test'
        )
        assert source == 'data/test.xlsx'
        assert fmt == 'excel'
        assert kwargs == {'table': 'Sheet1', 'skiprows': 'Year'}

    def test_file_dict_with_explicit_format(self):
        """file: dict with explicit format overrides auto-detection"""
        from pandoc_embedz.data_loader import _normalize_data_source
        source, fmt, kwargs = _normalize_data_source(
            {'file': 'data.dat', 'format': 'tsv'},
            'test'
        )
        assert fmt == 'tsv'
        assert kwargs == {}

    def test_file_dict_with_data_format_fallback(self):
        """file: dict uses data_format when no format in dict or extension"""
        from pandoc_embedz.data_loader import _normalize_data_source
        source, fmt, kwargs = _normalize_data_source(
            {'file': 'data.dat'},
            'test',
            data_format='json'
        )
        assert fmt == 'json'

    def test_dict_without_data_or_file_raises(self):
        """Dict without 'data' or 'file' key raises ValueError"""
        from pandoc_embedz.data_loader import _normalize_data_source
        with pytest.raises(ValueError, match="must have 'data' or 'file' key"):
            _normalize_data_source({'table': 'Sheet1'}, 'test')

    def test_multiline_string_returns_3_tuple(self):
        """Multiline string returns (StringIO, csv, {})"""
        from pandoc_embedz.data_loader import _normalize_data_source
        source, fmt, kwargs = _normalize_data_source('a,b\n1,2', 'test')
        assert isinstance(source, StringIO)
        assert fmt == 'csv'
        assert kwargs == {}


class TestIsResolvedData:
    """Tests for _is_resolved_data with file: dict"""

    def test_list_is_resolved(self):
        from pandoc_embedz.data_loader import _is_resolved_data
        assert _is_resolved_data([1, 2, 3]) is True

    def test_plain_dict_is_resolved(self):
        from pandoc_embedz.data_loader import _is_resolved_data
        assert _is_resolved_data({'key': 'value'}) is True

    def test_inline_data_dict_is_not_resolved(self):
        from pandoc_embedz.data_loader import _is_resolved_data
        assert _is_resolved_data({'data': 'a,b\n1,2'}) is False

    def test_format_dict_is_not_resolved(self):
        from pandoc_embedz.data_loader import _is_resolved_data
        assert _is_resolved_data({'format': 'csv'}) is False

    def test_file_dict_is_not_resolved(self):
        from pandoc_embedz.data_loader import _is_resolved_data
        assert _is_resolved_data({'file': 'data.xlsx', 'table': 'Sheet1'}) is False


class TestMultiTableExcel:
    """Tests for multi-table SQL with Excel files"""

    @pytest.fixture(autouse=True)
    def check_openpyxl(self):
        pytest.importorskip("openpyxl")

    def test_multi_table_excel_with_query(self, tmp_path):
        """Multi-table SQL query with Excel data sources"""
        import openpyxl
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        # Create Excel file with two sheets
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'products'
        ws1.append(['product_id', 'name', 'price'])
        ws1.append([1, 'Widget', 1280])
        ws1.append([2, 'Gadget', 2480])

        ws2 = wb.create_sheet('sales')
        ws2.append(['product_id', 'quantity'])
        ws2.append([1, 5])
        ws2.append([2, 3])

        path = str(tmp_path / 'multi.xlsx')
        wb.save(path)

        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        code = f'''---
data:
  products:
    file: {path}
    table: products
  sales:
    file: {path}
    table: sales
query: |
  SELECT p.name, p.price, s.quantity
  FROM sales s
  JOIN products p ON s.product_id = p.product_id
  ORDER BY p.name
---
{{% for row in data %}}
- {{{{ row.name }}}}: {{{{ row.quantity }}}} units at ¥{{{{ row.price }}}}
{{% endfor %}}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()
        result = process_embedz(elem, doc)

        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        assert 'Gadget' in markdown
        assert 'Widget' in markdown
        assert '5 units' in markdown
        assert '3 units' in markdown

    def test_multi_table_excel_with_skiprows(self, tmp_path):
        """Multi-table SQL with skiprows parameter in file: dict"""
        import openpyxl
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'data'
        ws.append(['Report Title'])
        ws.append(['Generated 2024'])
        ws.append([])
        ws.append(['name', 'value'])
        ws.append(['Arthur', 42])
        ws.append(['Ford', 100])

        path = str(tmp_path / 'skiprows.xlsx')
        wb.save(path)

        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        code = f'''---
data:
  items:
    file: {path}
    table: data
    skiprows: name
query: |
  SELECT * FROM items ORDER BY value DESC
---
{{% for row in data %}}
- {{{{ row.name }}}}: {{{{ row.value }}}}
{{% endfor %}}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()
        result = process_embedz(elem, doc)

        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        assert 'Ford' in markdown
        assert 'Arthur' in markdown

    def test_multi_table_file_dict_without_query(self, tmp_path):
        """Multi-table with file: dict but no query (direct access)"""
        import openpyxl
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'Sheet1'
        ws1.append(['name', 'value'])
        ws1.append(['Arthur', 42])
        ws2 = wb.create_sheet('Sheet2')
        ws2.append(['key', 'desc'])
        ws2.append(['author', 'Douglas Adams'])

        path = str(tmp_path / 'direct.xlsx')
        wb.save(path)

        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        code = f'''---
data:
  products:
    file: {path}
    table: Sheet1
  meta:
    file: {path}
    table: Sheet2
---
{{% for p in data.products %}}
- {{{{ p.name }}}}: {{{{ p.value }}}}
{{% endfor %}}
Author: {{{{ data.meta[0].desc }}}}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()
        result = process_embedz(elem, doc)

        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        assert 'Arthur: 42' in markdown
        assert 'Douglas Adams' in markdown

    def test_multi_table_mixed_file_dict_and_string(self, tmp_path):
        """Mix file: dict (Excel) with plain string (CSV)"""
        import openpyxl
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'items'
        ws.append(['id', 'name'])
        ws.append([1, 'Widget'])
        ws.append([2, 'Gadget'])

        xlsx_path = str(tmp_path / 'items.xlsx')
        wb.save(xlsx_path)

        csv_path = str(tmp_path / 'prices.csv')
        Path(csv_path).write_text('id,price\n1,1280\n2,2480\n')

        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        code = f'''---
data:
  items:
    file: {xlsx_path}
    table: items
  prices: {csv_path}
query: |
  SELECT i.name, p.price
  FROM items i JOIN prices p ON i.id = p.id
  ORDER BY i.name
---
{{% for row in data %}}
- {{{{ row.name }}}}: ¥{{{{ row.price }}}}
{{% endfor %}}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()
        result = process_embedz(elem, doc)

        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        assert 'Gadget' in markdown
        assert 'Widget' in markdown
        assert '1280' in markdown or '1,280' in markdown


class TestStartrow:
    """Tests for startrow parameter (replaces deprecated skiprows in filter.py)"""

    @pytest.fixture(autouse=True)
    def check_openpyxl(self):
        pytest.importorskip("openpyxl")

    def _make_excel_with_title(self, tmp_path):
        """Helper: create Excel with 2 title rows + header + data"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Annual Report 2024'])
        ws.append(['Generated: 2024-04-01'])
        ws.append([])
        ws.append(['name', 'value'])
        ws.append(['Arthur', 42])
        ws.append(['Ford', 100])
        path = str(tmp_path / 'startrow.xlsx')
        wb.save(path)
        return path

    def test_startrow_integer(self, tmp_path):
        """startrow integer is 1-indexed: startrow=3 skips 2 rows"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        path = self._make_excel_with_title(tmp_path)

        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        code = f'''---
data: {path}
startrow: 3
---
{{% for row in data %}}
- {{{{ row.name }}}}: {{{{ row.value }}}}
{{% endfor %}}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()
        result = process_embedz(elem, doc)

        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        assert 'Arthur: 42' in markdown
        assert 'Ford: 100' in markdown

    def test_startrow_string(self, tmp_path):
        """startrow string passes through as-is (same as skiprows)"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        path = self._make_excel_with_title(tmp_path)

        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        code = f'''---
data: {path}
startrow: name
---
{{% for row in data %}}
- {{{{ row.name }}}}: {{{{ row.value }}}}
{{% endfor %}}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()
        result = process_embedz(elem, doc)

        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        assert 'Arthur: 42' in markdown
        assert 'Ford: 100' in markdown

    def test_startrow_list(self, tmp_path):
        """startrow list passes through as-is (same as skiprows)"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Title', 'Report'])
        ws.append(['年', '月', '件数'])
        ws.append([2024, 4, 100])
        ws.append([2024, 5, 200])
        path = str(tmp_path / 'startrow_list.xlsx')
        wb.save(path)

        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        code = f'''---
data: {path}
startrow: [年, 月]
---
{{% for row in data %}}
- {{{{ row.年 }}}}-{{{{ row.月 }}}}: {{{{ row.件数 }}}}
{{% endfor %}}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()
        result = process_embedz(elem, doc)

        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        assert '2024-4: 100' in markdown
        assert '2024-5: 200' in markdown

    def test_skiprows_deprecation_warning(self, tmp_path, capsys):
        """skiprows still works but emits a deprecation warning"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        path = self._make_excel_with_title(tmp_path)

        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        code = f'''---
data: {path}
skiprows: 2
---
{{% for row in data %}}
- {{{{ row.name }}}}: {{{{ row.value }}}}
{{% endfor %}}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()
        result = process_embedz(elem, doc)

        captured = capsys.readouterr()
        assert 'deprecated' in captured.err.lower()

        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        assert 'Arthur: 42' in markdown

    def test_startrow_and_skiprows_both_error(self, tmp_path):
        """Specifying both startrow and skiprows raises ValueError"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        path = self._make_excel_with_title(tmp_path)

        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        code = f'''---
data: {path}
startrow: 3
skiprows: 2
---
{{% for row in data %}}
- {{{{ row.name }}}}
{{% endfor %}}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()

        with pytest.raises(ValueError, match="Cannot specify both"):
            process_embedz(elem, doc)

    def test_startrow_zero_raises_error(self, tmp_path):
        """startrow: 0 is invalid (must be >= 1)"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf

        path = self._make_excel_with_title(tmp_path)

        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        code = f'''---
data: {path}
startrow: 0
---
{{% for row in data %}}
- {{{{ row.name }}}}
{{% endfor %}}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()

        with pytest.raises(ValueError, match="must be >= 1"):
            process_embedz(elem, doc)

    def test_startrow_in_file_dict(self, tmp_path):
        """startrow works in file: dict syntax (multi-table)"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'data'
        ws.append(['Report Title'])
        ws.append(['Generated 2024'])
        ws.append([])
        ws.append(['name', 'value'])
        ws.append(['Arthur', 42])
        ws.append(['Ford', 100])
        path = str(tmp_path / 'file_dict.xlsx')
        wb.save(path)

        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        code = f'''---
data:
  items:
    file: {path}
    table: data
    startrow: name
query: |
  SELECT * FROM items ORDER BY value DESC
---
{{% for row in data %}}
- {{{{ row.name }}}}: {{{{ row.value }}}}
{{% endfor %}}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()
        result = process_embedz(elem, doc)

        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        assert 'Ford: 100' in markdown
        assert 'Arthur: 42' in markdown

    def test_startrow_integer_in_file_dict(self, tmp_path):
        """startrow integer conversion works in file: dict syntax"""
        from pandoc_embedz.filter import process_embedz, GLOBAL_VARS
        from pandoc_embedz.config import SAVED_TEMPLATES
        import panflute as pf
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'data'
        ws.append(['Report Title'])
        ws.append(['Generated 2024'])
        ws.append([])
        ws.append(['name', 'value'])
        ws.append(['Arthur', 42])
        ws.append(['Ford', 100])
        path = str(tmp_path / 'file_dict_int.xlsx')
        wb.save(path)

        SAVED_TEMPLATES.clear()
        GLOBAL_VARS.clear()

        code = f'''---
data:
  items:
    file: {path}
    table: data
    startrow: 3
query: |
  SELECT * FROM items ORDER BY value DESC
---
{{% for row in data %}}
- {{{{ row.name }}}}: {{{{ row.value }}}}
{{% endfor %}}'''

        elem = pf.CodeBlock(code, classes=['embedz'])
        doc = pf.Doc()
        result = process_embedz(elem, doc)

        if isinstance(result, list):
            markdown = pf.convert_text(result, input_format='panflute', output_format='markdown')
        else:
            markdown = pf.convert_text([result], input_format='panflute', output_format='markdown')

        assert 'Ford: 100' in markdown
        assert 'Arthur: 42' in markdown
